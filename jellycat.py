from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.touch_actions import TouchActions
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from fake_useragent import UserAgent
import time
import openpyxl
import math





def configure_driver():
    # Add additional Options to the webdriver
    chrome_options = Options()
    ua = UserAgent()
    userAgent = ua.random                                     #THIS IS FAKE AGENT IT WILL GIVE YOU NEW AGENT EVERYTIME
    print(userAgent)
   # add the argument and make the browser Headless.
   # chrome_options.add_argument("--headless")                    if you don't want to see the display on chrome just uncomment this
    chrome_options.add_argument(f'user-agent={userAgent}')     #useragent added
    chrome_options.add_argument("--log-level=3")               #removes error/warning/info messages displayed on the console
    chrome_options.add_argument("--disable-notifications")     #disable notifications
    chrome_options.add_argument("--disable-infobars")         #disable infobars ""Chrome is being controlled by automated test software"  Although is isn't supported by Chrome anymore
    chrome_options.add_argument("start-maximized")            #will maximize chrome screen
    chrome_options.add_argument('--disable-gpu')             #disable gpu (not load pictures fully)
    chrome_options.add_argument("--disable-extensions")       #will disable developer mode extensions
    #chrome_options.add_argument('--proxy-server=%s' % PROXY)
    #chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    #prefs = {"profile.managed_default_content_settings.images": 2}
    #chrome_options.add_experimental_option("prefs", prefs)             #we have disabled pictures (so no time is wasted in loading them)
    driver = webdriver.Chrome(ChromeDriverManager().install(), options = chrome_options)   #you don't have to download chromedriver it will be downloaded by itself and will be saved in cache
    return driver

def RunScrapper(driver):

    start_time = time.time()


    #workbook created
    wb = openpyxl.Workbook()
    # add_sheet is used to create sheet.
    sheet1 = wb.active
    print(" WORKSHEET CREATED SUCCESSFULLY!")
    # INITIALIZING THE COLOUMN NAMES NOW
    c1 = sheet1.cell(row = 1, column = 1)
    c1.value= "URL"
    c2 = sheet1.cell(row=1, column=2)
    c2.value = "Short Description"
    c3 = sheet1.cell(row=1, column=3)
    c3.value = "Inventory Status"
    c4 = sheet1.cell(row=1, column=4)
    c4.value = "Inventory Description"
    c5 = sheet1.cell(row=1, column=5)
    c5.value = "Long Description"
    c6 = sheet1.cell(row=1, column=6)
    c6.value = "Safety & Care"
    c7 = sheet1.cell(row=1, column=7)
    c7.value = "Navigation Path"
    c8 = sheet1.cell(row=1, column=8)
    c8.value = "Product Name"
    c9 = sheet1.cell(row=1, column=9)
    c9.value = "Size"
    c10 = sheet1.cell(row=1, column=10)
    c10.value = "Color"
    c11 = sheet1.cell(row=1, column=11)
    c11.value = "Regular Price"
    c12 = sheet1.cell(row=1, column=12)
    c12.value = "Sale Price"
    c13 = sheet1.cell(row=1, column=13)
    c13.value = "SKU"
    c14 = sheet1.cell(row=1, column=14)
    c14.value = "Images"
    wb.save("jellycat.xlsx")
    #setting row number to 2
    mi=2


    mainlink="https://www.jellycat.com/eu/"
    driver.get(mainlink)
    WebDriverWait(driver,40).until(expected_conditions.visibility_of_element_located((By.ID,'nav-level0')))
    try:
        driver.find_element_by_xpath('//input[@value="Close"]').click()
    except Exception:
        pass
    navbar=driver.find_element_by_id('nav-level0')
    links=navbar.find_elements_by_tag_name('a')
    x=1
    majlinks=[]
    for l in links:
       if x==1:
           pass
       else:
           majlinks.append(l.get_attribute('href'))
       x+=1
    alllinks=[]
    for m in majlinks:
        driver.get(m)
        WebDriverWait(driver,40).until(expected_conditions.visibility_of_all_elements_located((By.XPATH,'//div[@class="mtb0-5"]')))
        container=driver.find_elements_by_xpath('//div[@class="mtb0-5"]')
        for contain in container:
            link=contain.find_element_by_tag_name('a').get_attribute('href')
            alllinks.append(link)
    entno = 1
    catno=1
    for al in alllinks:
        print("Category: ",al)
        print("CategoryNo: ",catno)
        catno+=1
        try:
            driver.get(al)
            linkstogo = []
            WebDriverWait(driver, 40).until(
                expected_conditions.visibility_of_element_located((By.ID, 'productDataOnPage')))
            noitems = driver.find_element_by_id('productDataNavCTOP').text
            noitems = noitems.replace("items", "")
            noitems = noitems.strip()
            noitems = int(noitems)
            if noitems >= 24:
                scrollno = int(math.ceil(noitems / 24))
                for sn in range(scrollno):
                    element = driver.find_element_by_id("listing-footer")
                    actions = ActionChains(driver)
                    actions.move_to_element(element).perform()
                    time.sleep(3)
            else:
                pass
            WebDriverWait(driver, 40).until(
                expected_conditions.visibility_of_element_located((By.ID, 'productDataOnPage')))
            mainbl = driver.find_element_by_id('productDataOnPage')
            container = mainbl.find_elements_by_xpath('//a[@data-listing="name"]')
            print("Total Products In This Page: ", len(container))
            for contain in container:
                linkstogo.append(contain.get_attribute('href'))

            for linktogo in linkstogo:
                print("Entry No: ", entno)
                entno += 1
                print("SCRAPING: ", linktogo)
                try:
                    driver.get(linktogo)

                    WebDriverWait(driver, 40).until(
                        expected_conditions.visibility_of_element_located((By.XPATH, '//h1[@class="mtb0-5 f-capi"]')))

                    navigationpath = driver.find_element_by_xpath('//div[@class="f-color2 f-brand-persist-links"]').text

                    try:
                        shortdescription = driver.find_element_by_xpath('//div[@class="mtb0-5"]').text
                    except Exception:
                        shortdescription = ''
                        pass
                    try:
                        longdescription = driver.find_element_by_xpath(
                            '//div[@class="productbody no-gaps f-xspace"]').text
                    except Exception:
                        longdescription = ''
                        pass

                    try:
                        pd = driver.find_element_by_id('ProductDetails')
                        info = pd.find_element_by_xpath('.//div[@class="accordion f-13"]')
                        heading = info.find_elements_by_tag_name('h3')
                        headingbody = info.find_elements_by_tag_name('div')
                        hbc = 0
                        safetycare = ''
                        try:
                            for hd in range(len(heading)):
                                head = heading[hd].text
                                if "SAFETY & CARE" in head:
                                    st = headingbody[hbc].get_attribute('style')
                                    if "display: none;" in st:
                                        try:
                                            heading[hd].click()
                                        except Exception:
                                            pass
                                        safetycare = headingbody[hbc].text
                                    else:
                                        safetycare = headingbody[hbc].text
                                    break
                                else:
                                    pass
                                hbc += 2
                        except Exception:
                            safetycare = ''
                            pass
                    except Exception:
                        safetycare = ''
                        pass

                    print("Short Description: ", shortdescription)
                    print("Long Description: ", longdescription)
                    print("Safety & Care: ", safetycare)
                    print("Navigation: ", navigationpath)

                    colormain1 = driver.find_element_by_xpath('//div[@class="nogaps"]')
                    colors1 = colormain1.find_elements_by_xpath(
                        './/div[@class="pointer width4 height4 inline-block mr0-5 mb0-5"]')

                    for color in range(len(colors1)):
                        colormain = driver.find_element_by_xpath('//div[@class="nogaps"]')
                        colors = colormain.find_elements_by_xpath(
                            './/div[@class="pointer width4 height4 inline-block mr0-5 mb0-5"]')
                        colorname = colors[color].find_element_by_tag_name('img').get_attribute('alt')
                        colors[color].click()
                        sizemain1 = driver.find_element_by_xpath('//div[@class="f-13 nogaps"]')
                        sizes1 = sizemain1.find_elements_by_xpath(
                            './/div[@class="pointer width6 height6 inline-block mr0-5 mb0-5 f-upper"]')
                        for size in range(len(sizes1)):
                            sizemain = driver.find_element_by_xpath('//div[@class="f-13 nogaps"]')
                            sizes = sizemain.find_elements_by_xpath(
                                './/div[@class="pointer width6 height6 inline-block mr0-5 mb0-5 f-upper"]')
                            sizename = sizes[size].text
                            sizes[size].click()

                            productname = driver.find_element_by_xpath('//h1[@data-bind="text:name"]').text

                            sku = driver.find_element_by_xpath('//div[@data-bind="text:sku"]').text

                            pricexpath = '//span[@data-bind="text: price, css:pricecss, style:{' + "'display'" + ':pricedisplay}"]'
                            price = driver.find_element_by_xpath(pricexpath).text
                            regularprice = price
                            saleprice = price

                            allimages = ''
                            try:
                                altimgs = driver.find_element_by_id('alternativeImages')
                                imgs = altimgs.find_elements_by_tag_name('img')
                                x = 1
                                for img in imgs:
                                    if x == 1:
                                        allimages = img.get_attribute('src')
                                    else:
                                        allimages = allimages + ", " + img.get_attribute('src')
                                    x += 1
                            except Exception:
                                allimages = ''
                                pass

                            try:
                                inventorystatus = driver.find_element_by_xpath(
                                    '//form[@class="fieldwithbutton mb"]').text
                                inventorystatus=inventorystatus.strip()
                                if inventorystatus=='':
                                    inventorystatus='IN-STOCK'
                                else:
                                    pass
                            except Exception:
                                inventorystatus = ''
                                pass

                            try:
                                inventorydescription = driver.find_element_by_xpath('//div[@class="mt0-25"]').text
                            except Exception:
                                inventorydescription = ''
                                pass

                            print("Product Name: ", productname)
                            print("Size: ", sizename)
                            print("Color Name: ", colorname)
                            print("SKU: ", sku)
                            print("Price: ", price)
                            print("Images: ", allimages)
                            print("Inventory Status: ", inventorystatus)
                            print("Inventory Description: ", inventorydescription)
                            c1 = sheet1.cell(row=mi, column=1)
                            c1.value = linktogo
                            c2 = sheet1.cell(row=mi, column=2)
                            c2.value = shortdescription
                            c3 = sheet1.cell(row=mi, column=3)
                            c3.value = inventorystatus
                            c4 = sheet1.cell(row=mi, column=4)
                            c4.value = inventorydescription
                            c5 = sheet1.cell(row=mi, column=5)
                            c5.value = longdescription
                            c6 = sheet1.cell(row=mi, column=6)
                            c6.value = safetycare
                            c7 = sheet1.cell(row=mi, column=7)
                            c7.value = navigationpath
                            c8 = sheet1.cell(row=mi, column=8)
                            c8.value = productname
                            c9 = sheet1.cell(row=mi, column=9)
                            c9.value = sizename
                            c10 = sheet1.cell(row=mi, column=10)
                            c10.value = colorname
                            c11 = sheet1.cell(row=mi, column=11)
                            c11.value = regularprice
                            c12 = sheet1.cell(row=mi, column=12)
                            c12.value = saleprice
                            c13 = sheet1.cell(row=mi, column=13)
                            c13.value = sku
                            c14 = sheet1.cell(row=mi, column=14)
                            c14.value = allimages
                            mi += 1
                            print("---------------------")
                    wb.save("jellycat.xlsx")
                except Exception:
                    print("Product Scrapping Failed")
                    pass
                print("")
                print("**************************************************************")
                print("")
        except Exception:
            print("Broken Category")
            pass
        print("")
        print("**************************************************************")
        print("")
        print("")
        print("**************************************************************")
        print("")

    #give time taken to execute everything
    print("time elapsed: {:.2f}s".format(time.time() - start_time))



# create the driver object.
driver= configure_driver()

#call the scrapper to run
RunScrapper(driver)

# close the driver.
#driver.close()














